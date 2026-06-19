#!/usr/bin/env python3
"""
Generates "Goal Tracker Template.xlsx" -- an Excel companion to the
Discipline web dashboard (index.html), styled as a warm cream/taupe
"habit tracker" sheet: a daily checkbox grid grouped into weeks, a
per-habit Goal/Actual/Left/Progress analysis table, a habit ranking
list, Goal/Completed/Left summary cards with an overall-status donut
chart, Daily/Weekly progress bar charts, and a Mood/Motivation tracker
with a line chart.

Everything lives on a single "Tracker" sheet, sized and zoomed to fit
one screen with no horizontal scrolling. A hidden "Helpers" sheet holds
the streak/weekly/ranking/donut calculations behind it.

Edit HABITS (and EMOJI) below and re-run this script to regenerate.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from datetime import date

# ---------------------------------------------------------------------------
# Editable habit list
# ---------------------------------------------------------------------------
HABITS = [
    "Wake Up at 6:00 AM",
    "Meditation / Stretch",
    "Hit the Gym",
    "Cold Shower",
    "Read 10 Pages",
    "Pray",
    "Deep Work",
    "Limit Social Media",
]
EMOJI = ["⏰", "\U0001F9D8", "\U0001F4AA", "\U0001F6BF", "\U0001F4D6", "\U0001F64F", "\U0001F4BB", "\U0001F4F5"]

DAY_COLS = 31
CHECK = "✓"

# ---------------------------------------------------------------------------
# Palette -- warm cream/taupe "aesthetic habit tracker" look
# ---------------------------------------------------------------------------
CREAM = "FAF6EF"        # page background
PANEL = "F1E9DC"        # subtle panel fill
DARK = "6E5F52"         # taupe header/section bars
DARK_TEXT = "FAF6EF"    # text on dark bars
INK = "4A3F35"          # primary body text
MUTED = "9C8F80"        # secondary text
ACCENT = "B5694A"       # terracotta accent
BORDER_GRAY = "DCD3C4"

CHECK_FILL_COLOR = "E3D9C8"
CHECK_FONT_COLOR = "5C4A3A"

TIER_HIGH_FILL = "DCE5D0"
TIER_HIGH_FONT = "55703F"
TIER_MID_FILL = "F2E2C4"
TIER_MID_FONT = "8A6A33"
TIER_LOW_FILL = "F1DAD3"
TIER_LOW_FONT = "98503F"

DATABAR_COLOR = "C2A98A"

CREAM_FILL = PatternFill("solid", fgColor=CREAM)
PANEL_FILL = PatternFill("solid", fgColor=PANEL)
HEADER_FILL = PatternFill("solid", fgColor=DARK)
HEADER_FONT = Font(color=DARK_TEXT, bold=True, size=10, name="Segoe UI")
SECTION_FONT = Font(color=DARK_TEXT, bold=True, size=11, name="Segoe UI")
TITLE_FONT = Font(color=ACCENT, bold=True, size=20, name="Segoe UI")
SUBTITLE_FONT = Font(color=MUTED, bold=True, size=11, name="Segoe UI", italic=True)
LABEL_FONT = Font(color=INK, bold=True, size=10, name="Segoe UI")
MUTED_FONT = Font(color=MUTED, size=9, name="Segoe UI")
CELL_FONT = Font(color=INK, size=10, name="Segoe UI")
STAT_LABEL_FONT = Font(color=MUTED, bold=True, size=9, name="Segoe UI")
STAT_VALUE_FONT = Font(color=ACCENT, bold=True, size=18, name="Segoe UI")

THIN = Side(style="thin", color=BORDER_GRAY)
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM = Side(style="medium", color="C7B9A6")
BOX_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

NA_FILL = PatternFill("solid", fgColor=PANEL)

CHECK_FILL = PatternFill("solid", fgColor=CHECK_FILL_COLOR)
CHECKBOX_FONT = Font(color=INK, size=14, name="Segoe UI")
CHECK_FONT_OBJ = Font(color=CHECK_FONT_COLOR, bold=True, size=14, name="Segoe UI")
TIER_HIGH_FILL_OBJ = PatternFill("solid", fgColor=TIER_HIGH_FILL)
TIER_HIGH_FONT_OBJ = Font(color=TIER_HIGH_FONT, bold=True, size=10, name="Segoe UI")
TIER_MID_FILL_OBJ = PatternFill("solid", fgColor=TIER_MID_FILL)
TIER_MID_FONT_OBJ = Font(color=TIER_MID_FONT, bold=True, size=10, name="Segoe UI")
TIER_LOW_FILL_OBJ = PatternFill("solid", fgColor=TIER_LOW_FILL)
TIER_LOW_FONT_OBJ = Font(color=TIER_LOW_FONT, bold=True, size=10, name="Segoe UI")


def col(n):
    """1-based column index -> letter."""
    return get_column_letter(n)


# ---------------------------------------------------------------------------
# Layout constants (single "Tracker" sheet)
# ---------------------------------------------------------------------------
N = len(HABITS)

COL_HABIT = 1                                  # A
FIRST_DAY_COL = 2                              # B
LAST_DAY_COL = FIRST_DAY_COL + DAY_COLS - 1    # 32 = AF
SPACER1_COL = LAST_DAY_COL + 1                 # 33 = AG
ANALYSIS_FIRST = SPACER1_COL + 1               # 34 = AH
ANALYSIS_HEADERS = ["GOAL", "ACTUAL", "LEFT", "PROGRESS", "STREAK", "BEST"]
ANALYSIS_LAST = ANALYSIS_FIRST + len(ANALYSIS_HEADERS) - 1   # 39 = AM
SPACER2_COL = ANALYSIS_LAST + 1                # 40 = AN
RANK_NUM_COL = SPACER2_COL + 1                 # 41 = AO
RANK_NAME_COL = RANK_NUM_COL + 1               # 42 = AP
LAST_COL = RANK_NAME_COL

ZOOM_SCALE = 75

TITLE_ROW = 1
SUBTITLE_ROW = 2
CONFIG_HEADER_ROW = 4
CONFIG_ROW1 = 5
CONFIG_ROW2 = 6

DAILY_SECTION_ROW = 14
GRID_WEEK_ROW = 15
WEEKDAY_ROW = 16
DAYNUM_ROW = 17
HABIT_FIRST_ROW = 18
HABIT_LAST_ROW = HABIT_FIRST_ROW + N - 1       # 25
SCORE_ROW = HABIT_LAST_ROW + 1                 # 26
MOOD_SECTION_ROW = SCORE_ROW + 1               # 27
MOOD_ROW = MOOD_SECTION_ROW + 1                # 28
MOTIVATION_ROW = MOOD_ROW + 1                  # 29
LINE_CHART_ROW = MOTIVATION_ROW + 2            # 31

# Week groupings (1-indexed column ranges within the day grid)
WEEK_GROUPS = [
    (FIRST_DAY_COL + 0, FIRST_DAY_COL + 6),    # days 1-7
    (FIRST_DAY_COL + 7, FIRST_DAY_COL + 13),   # days 8-14
    (FIRST_DAY_COL + 14, FIRST_DAY_COL + 20),  # days 15-21
    (FIRST_DAY_COL + 21, FIRST_DAY_COL + 27),  # days 22-28
    (FIRST_DAY_COL + 28, LAST_DAY_COL),        # days 29-31
]

# ---------------------------------------------------------------------------
# Helpers-sheet layout
# ---------------------------------------------------------------------------
HELPER_STREAK_FIRST_ROW = 2        # rows 2..(2+N-1): per-habit streak running counts
WEEK_LABEL_ROW = 12                # A12:E12 = "Week 1".."Week 5"
WEEK_VALUE_ROW = 13                # A13:E13 = weekly score %
DONUT_FIRST_ROW = 15               # rows 15-16: Completed / Remaining
TIEBREAK_FIRST_ROW = 18            # rows 18..(18+N-1): ranking tie-break values


def section_bar(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = HEADER_FILL
    c.font = SECTION_FONT
    c.alignment = LEFT
    ws.row_dimensions[row].height = 20


def build_tracker(wb):
    ws = wb.create_sheet("Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = ZOOM_SCALE

    # Cream page background as a base layer
    for r in range(1, LINE_CHART_ROW + 16):
        for c in range(1, LAST_COL + 1):
            ws.cell(row=r, column=c).fill = CREAM_FILL

    last_day_letter = col(LAST_DAY_COL)

    # --- Title ---
    ws.merge_cells(start_row=TITLE_ROW, start_column=1, end_row=TITLE_ROW, end_column=5)
    title = ws.cell(row=TITLE_ROW, column=1, value="HABIT TRACKER")
    title.font = TITLE_FONT
    title.alignment = LEFT
    ws.row_dimensions[TITLE_ROW].height = 30

    ws.merge_cells(start_row=SUBTITLE_ROW, start_column=1, end_row=SUBTITLE_ROW, end_column=5)
    subtitle = ws.cell(
        row=SUBTITLE_ROW, column=1,
        value='="- "&TEXT(DATE($D$5,$B$5,1),"mmmm")&" "&$D$5&" -"'
    )
    subtitle.font = SUBTITLE_FONT
    subtitle.alignment = LEFT

    # --- Calendar settings ---
    ws.merge_cells(start_row=CONFIG_HEADER_ROW, start_column=1, end_row=CONFIG_HEADER_ROW, end_column=5)
    cfg = ws.cell(row=CONFIG_HEADER_ROW, column=1, value="CALENDAR SETTINGS")
    cfg.fill = HEADER_FILL
    cfg.font = SECTION_FONT
    cfg.alignment = LEFT
    ws.row_dimensions[CONFIG_HEADER_ROW].height = 20

    today = date.today()

    ws.cell(row=CONFIG_ROW1, column=1, value="Month (1-12):").font = LABEL_FONT
    ws.cell(row=CONFIG_ROW1, column=1).alignment = LEFT
    b5 = ws.cell(row=CONFIG_ROW1, column=2, value=today.month)
    b5.font = CELL_FONT
    b5.alignment = CENTER
    b5.fill = PANEL_FILL
    b5.border = BORDER_ALL
    dv_month = DataValidation(type="whole", operator="between", formula1=1, formula2=12)
    ws.add_data_validation(dv_month)
    dv_month.add(b5)

    ws.cell(row=CONFIG_ROW1, column=4, value="Year:").font = LABEL_FONT
    ws.cell(row=CONFIG_ROW1, column=4).alignment = LEFT
    d5 = ws.cell(row=CONFIG_ROW1, column=5, value=today.year)
    d5.font = CELL_FONT
    d5.alignment = CENTER
    d5.fill = PANEL_FILL
    d5.border = BORDER_ALL

    ws.cell(row=CONFIG_ROW2, column=1, value="Days in Month:").font = LABEL_FONT
    ws.cell(row=CONFIG_ROW2, column=1).alignment = LEFT
    b6 = ws.cell(row=CONFIG_ROW2, column=2, value="=DAY(EOMONTH(DATE($D$5,$B$5,1),0))")
    b6.font = MUTED_FONT
    b6.alignment = CENTER
    b6.fill = PANEL_FILL
    b6.border = BORDER_ALL

    ws.cell(row=CONFIG_ROW2, column=4, value="Today (day #):").font = LABEL_FONT
    ws.cell(row=CONFIG_ROW2, column=4).alignment = LEFT
    d6 = ws.cell(row=CONFIG_ROW2, column=5,
                  value='=IF(AND(YEAR(TODAY())=$D$5,MONTH(TODAY())=$B$5),DAY(TODAY()),$B$6)')
    d6.font = MUTED_FONT
    d6.alignment = CENTER
    d6.fill = PANEL_FILL
    d6.border = BORDER_ALL

    # --- Stat cards: Goal / Completed / Left ---
    stat_rows = [
        (1, "GOAL", f"={N}*$B$6"),
        (2, "COMPLETED", "=Helpers!B15"),
        (3, "LEFT", "=Helpers!B16"),
    ]
    for r, label, formula in stat_rows:
        lbl = ws.cell(row=r, column=ANALYSIS_FIRST, value=label)
        lbl.font = STAT_LABEL_FONT
        lbl.alignment = LEFT
        lbl.fill = PANEL_FILL
        lbl.border = BORDER_ALL
        ws.merge_cells(start_row=r, start_column=ANALYSIS_FIRST + 1, end_row=r, end_column=ANALYSIS_LAST)
        val = ws.cell(row=r, column=ANALYSIS_FIRST + 1, value=formula)
        val.font = STAT_VALUE_FONT
        val.alignment = Alignment(horizontal="right", vertical="center")
        val.fill = PANEL_FILL
        val.border = BORDER_ALL
        ws.row_dimensions[r].height = 22

    # --- Daily tracking grid ---
    section_bar(ws, DAILY_SECTION_ROW,
                 f"DAILY HABIT TRACKER  —  click a day's box and choose {CHECK} to mark it done")

    # "MY HABITS" label spanning the header rows
    ws.merge_cells(start_row=GRID_WEEK_ROW, start_column=COL_HABIT, end_row=DAYNUM_ROW, end_column=COL_HABIT)
    habits_lbl = ws.cell(row=GRID_WEEK_ROW, column=COL_HABIT, value="MY HABITS")
    habits_lbl.fill = HEADER_FILL
    habits_lbl.font = HEADER_FONT
    habits_lbl.alignment = WRAP_CENTER

    # Week group headers
    for i, (start, end) in enumerate(WEEK_GROUPS, start=1):
        ws.merge_cells(start_row=GRID_WEEK_ROW, start_column=start, end_row=GRID_WEEK_ROW, end_column=end)
        c = ws.cell(row=GRID_WEEK_ROW, column=start, value=f"Week {i}")
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        for cc in range(start, end + 1):
            ws.cell(row=GRID_WEEK_ROW, column=cc).border = BORDER_ALL
            ws.cell(row=GRID_WEEK_ROW, column=cc).fill = HEADER_FILL

    # "ANALYSIS" + "HABIT RANKING" group headers (rows GRID_WEEK_ROW..DAYNUM_ROW)
    ws.merge_cells(start_row=GRID_WEEK_ROW, start_column=ANALYSIS_FIRST, end_row=GRID_WEEK_ROW, end_column=ANALYSIS_LAST)
    a_hdr = ws.cell(row=GRID_WEEK_ROW, column=ANALYSIS_FIRST, value="ANALYSIS")
    a_hdr.fill = HEADER_FILL
    a_hdr.font = HEADER_FONT
    a_hdr.alignment = CENTER

    ws.merge_cells(start_row=GRID_WEEK_ROW, start_column=RANK_NUM_COL, end_row=DAYNUM_ROW, end_column=RANK_NAME_COL)
    r_hdr = ws.cell(row=GRID_WEEK_ROW, column=RANK_NUM_COL, value="HABIT RANKING")
    r_hdr.fill = HEADER_FILL
    r_hdr.font = HEADER_FONT
    r_hdr.alignment = WRAP_CENTER

    # Weekday letters row
    for d in range(1, DAY_COLS + 1):
        c_letter = col(FIRST_DAY_COL + d - 1)
        cell = ws.cell(row=WEEKDAY_ROW, column=FIRST_DAY_COL + d - 1)
        cell.value = (
            f'=IF({c_letter}{DAYNUM_ROW}<=$B$6,TEXT(DATE($D$5,$B$5,{c_letter}{DAYNUM_ROW}),"ddd"),"")'
        )
        cell.font = MUTED_FONT
        cell.alignment = CENTER
        cell.border = BORDER_ALL
        cell.fill = PANEL_FILL
    for c_idx in range(ANALYSIS_FIRST, ANALYSIS_LAST + 1):
        cell = ws.cell(row=WEEKDAY_ROW, column=c_idx)
        cell.fill = NA_FILL
        cell.border = BORDER_ALL

    # Day-number row + Analysis column headers
    for d in range(1, DAY_COLS + 1):
        c = ws.cell(row=DAYNUM_ROW, column=FIRST_DAY_COL + d - 1, value=d)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER_ALL
    for i, h in enumerate(ANALYSIS_HEADERS):
        c = ws.cell(row=DAYNUM_ROW, column=ANALYSIS_FIRST + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER_ALL

    # Habit rows
    for i, name in enumerate(HABITS):
        r = HABIT_FIRST_ROW + i
        ws.row_dimensions[r].height = 20
        name_cell = ws.cell(row=r, column=COL_HABIT, value=f"{EMOJI[i]}  {name}")
        name_cell.font = LABEL_FONT
        name_cell.alignment = LEFT
        name_cell.border = BORDER_ALL
        name_cell.fill = PANEL_FILL

        day_range = f"{col(FIRST_DAY_COL)}{r}:{last_day_letter}{r}"

        for d in range(1, DAY_COLS + 1):
            cell = ws.cell(row=r, column=FIRST_DAY_COL + d - 1)
            cell.alignment = CENTER
            cell.font = CHECKBOX_FONT
            cell.border = BOX_BORDER
            dv = DataValidation(type="list", formula1=f'"{CHECK},"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(cell)

        # Analysis table: Goal / Actual / Left / Progress / Streak / Best
        goal_c = ws.cell(row=r, column=ANALYSIS_FIRST, value="=$B$6")
        actual_c = ws.cell(row=r, column=ANALYSIS_FIRST + 1, value=f'=COUNTIF({day_range},"{CHECK}")')
        left_c = ws.cell(row=r, column=ANALYSIS_FIRST + 2,
                          value=f"={col(ANALYSIS_FIRST)}{r}-{col(ANALYSIS_FIRST + 1)}{r}")
        progress_c = ws.cell(
            row=r, column=ANALYSIS_FIRST + 3,
            value=(f"=IF({col(ANALYSIS_FIRST)}{r}=0,0,"
                   f"{col(ANALYSIS_FIRST + 1)}{r}/{col(ANALYSIS_FIRST)}{r})")
        )
        progress_c.number_format = "0%"

        helper_row = HELPER_STREAK_FIRST_ROW + i
        helper_range = f"Helpers!B{helper_row}:AF{helper_row}"
        streak_c = ws.cell(
            row=r, column=ANALYSIS_FIRST + 4,
            value=(f'=IF($D$6=1,INDEX({helper_range},1),'
                   f'MAX(INDEX({helper_range},$D$6),INDEX({helper_range},MAX(1,$D$6-1))))')
        )
        best_c = ws.cell(row=r, column=ANALYSIS_FIRST + 5, value=f"=MAX({helper_range})")

        for cell in (goal_c, actual_c, left_c, streak_c, best_c):
            cell.font = CELL_FONT
            cell.alignment = CENTER
            cell.border = BORDER_ALL
            cell.fill = PANEL_FILL
        progress_c.font = CELL_FONT
        progress_c.alignment = CENTER
        progress_c.border = BORDER_ALL
        progress_c.fill = PANEL_FILL

        # Habit ranking list (rank i+1)
        rank_num = ws.cell(row=r, column=RANK_NUM_COL, value=i + 1)
        rank_num.font = LABEL_FONT
        rank_num.alignment = CENTER
        rank_num.border = BORDER_ALL
        rank_num.fill = PANEL_FILL

        tie_range = f"Helpers!$B${TIEBREAK_FIRST_ROW}:$B${TIEBREAK_FIRST_ROW + N - 1}"
        match_expr = f"MATCH(LARGE({tie_range},{col(RANK_NUM_COL)}{r}),{tie_range},0)"
        progress_col_letter = col(ANALYSIS_FIRST + 3)
        rank_name = ws.cell(
            row=r, column=RANK_NAME_COL,
            value=(f'=INDEX($A${HABIT_FIRST_ROW}:$A${HABIT_LAST_ROW},{match_expr})'
                   f'&"  ("&TEXT(INDEX(${progress_col_letter}${HABIT_FIRST_ROW}:'
                   f'${progress_col_letter}${HABIT_LAST_ROW},{match_expr}),"0%")&")"')
        )
        rank_name.font = CELL_FONT
        rank_name.alignment = LEFT
        rank_name.border = BORDER_ALL
        rank_name.fill = PANEL_FILL

    # Conditional formatting: checkbox cells -> warm fill when checked
    check_range = f"{col(FIRST_DAY_COL)}{HABIT_FIRST_ROW}:{last_day_letter}{HABIT_LAST_ROW}"
    ws.conditional_formatting.add(
        check_range,
        CellIsRule(operator="equal", formula=[f'"{CHECK}"'], fill=CHECK_FILL, font=CHECK_FONT_OBJ),
    )

    # Conditional formatting: Progress column + Daily Score row -> tiers
    progress_range = f"{col(ANALYSIS_FIRST + 3)}{HABIT_FIRST_ROW}:{col(ANALYSIS_FIRST + 3)}{HABIT_LAST_ROW}"
    score_range = f"{col(FIRST_DAY_COL)}{SCORE_ROW}:{last_day_letter}{SCORE_ROW}"
    for rng in (progress_range, score_range):
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="greaterThanOrEqual", formula=["0.8"],
                              fill=TIER_HIGH_FILL_OBJ, font=TIER_HIGH_FONT_OBJ)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="between", formula=["0.5", "0.7999"],
                              fill=TIER_MID_FILL_OBJ, font=TIER_MID_FONT_OBJ)
        )
        ws.conditional_formatting.add(
            rng, CellIsRule(operator="lessThan", formula=["0.5"],
                              fill=TIER_LOW_FILL_OBJ, font=TIER_LOW_FONT_OBJ)
        )

    # Data bar for the Progress column
    ws.conditional_formatting.add(
        progress_range,
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                     color=DATABAR_COLOR, showValue=True)
    )

    # --- Daily score row ---
    label = ws.cell(row=SCORE_ROW, column=COL_HABIT, value="DAILY SCORE")
    label.font = LABEL_FONT
    label.border = BORDER_ALL
    label.fill = PANEL_FILL
    for d in range(1, DAY_COLS + 1):
        c_letter = col(FIRST_DAY_COL + d - 1)
        habit_range = f"{c_letter}{HABIT_FIRST_ROW}:{c_letter}{HABIT_LAST_ROW}"
        cell = ws.cell(row=SCORE_ROW, column=FIRST_DAY_COL + d - 1)
        cell.value = f'=IF({c_letter}{DAYNUM_ROW}<=$B$6,COUNTIF({habit_range},"{CHECK}")/{N},"")'
        cell.number_format = "0%"
        cell.font = CELL_FONT
        cell.alignment = CENTER
        cell.border = BORDER_ALL
        cell.fill = PANEL_FILL
    for c_idx in range(ANALYSIS_FIRST, LAST_COL + 1):
        cell = ws.cell(row=SCORE_ROW, column=c_idx)
        cell.border = BORDER_ALL
        cell.fill = NA_FILL

    # --- Mood / Motivation tracker ---
    section_bar(ws, MOOD_SECTION_ROW, "MENTAL STATE TRACKER  —  rate 1-10 each day")
    for row, text in ((MOOD_ROW, "MOOD"), (MOTIVATION_ROW, "MOTIVATION")):
        label = ws.cell(row=row, column=COL_HABIT, value=text)
        label.font = LABEL_FONT
        label.border = BORDER_ALL
        label.fill = PANEL_FILL
        for d in range(1, DAY_COLS + 1):
            cell = ws.cell(row=row, column=FIRST_DAY_COL + d - 1)
            cell.alignment = CENTER
            cell.font = CELL_FONT
            cell.border = BORDER_ALL
            cell.fill = PANEL_FILL
            dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(cell)
        for c_idx in range(ANALYSIS_FIRST, LAST_COL + 1):
            cell = ws.cell(row=row, column=c_idx)
            cell.border = BORDER_ALL
            cell.fill = NA_FILL

    build_charts(ws, wb)

    # --- Watermark / credit ---
    credit = ws.cell(row=47, column=1, value="© Nikolas Moore — Discipline Goal Tracker")
    credit.font = MUTED_FONT

    # --- Column widths ---
    ws.column_dimensions[col(COL_HABIT)].width = 26
    for d in range(1, DAY_COLS + 1):
        ws.column_dimensions[col(FIRST_DAY_COL + d - 1)].width = 3.0
    ws.column_dimensions[col(SPACER1_COL)].width = 1.5
    for i, w in enumerate([6, 6.5, 6, 10, 7, 6]):
        ws.column_dimensions[col(ANALYSIS_FIRST + i)].width = w
    ws.column_dimensions[col(SPACER2_COL)].width = 1.5
    ws.column_dimensions[col(RANK_NUM_COL)].width = 4
    ws.column_dimensions[col(RANK_NAME_COL)].width = 24

    ws.freeze_panes = "B18"
    return ws


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------
def build_charts(ws, wb):
    helpers_ws = wb["Helpers"]

    # Daily Progress (bar): Daily Score row across the day grid
    chart1 = BarChart()
    chart1.title = "Daily Progress"
    chart1.style = 10
    data1 = Reference(ws, min_col=FIRST_DAY_COL, max_col=LAST_DAY_COL, min_row=SCORE_ROW)
    cats1 = Reference(ws, min_col=FIRST_DAY_COL, max_col=LAST_DAY_COL, min_row=DAYNUM_ROW)
    chart1.add_data(data1, titles_from_data=False, from_rows=True)
    chart1.set_categories(cats1)
    chart1.y_axis.numFmt = "0%"
    chart1.height = 6
    chart1.width = 7.5
    chart1.legend = None
    ws.add_chart(chart1, "G1")

    # Weekly Progress (bar): from Helpers weekly score row
    chart2 = BarChart()
    chart2.title = "Weekly Progress"
    chart2.style = 10
    data2 = Reference(helpers_ws, min_col=1, max_col=5, min_row=WEEK_VALUE_ROW)
    cats2 = Reference(helpers_ws, min_col=1, max_col=5, min_row=WEEK_LABEL_ROW)
    chart2.add_data(data2, titles_from_data=False, from_rows=True)
    chart2.set_categories(cats2)
    chart2.y_axis.numFmt = "0%"
    chart2.height = 6
    chart2.width = 7.5
    chart2.legend = None
    ws.add_chart(chart2, "S1")

    # Overall Status (donut): Completed vs Remaining
    chart3 = DoughnutChart()
    chart3.title = "Overall Status"
    chart3.holeSize = 60
    data3 = Reference(helpers_ws, min_col=2, min_row=DONUT_FIRST_ROW, max_row=DONUT_FIRST_ROW + 1)
    cats3 = Reference(helpers_ws, min_col=1, min_row=DONUT_FIRST_ROW, max_row=DONUT_FIRST_ROW + 1)
    chart3.add_data(data3, titles_from_data=False)
    chart3.set_categories(cats3)
    chart3.dataLabels = DataLabelList()
    chart3.dataLabels.showPercent = True
    chart3.height = 5
    chart3.width = 6
    ws.add_chart(chart3, f"{col(ANALYSIS_FIRST)}4")

    # Mood / Motivation (line)
    chart4 = LineChart()
    chart4.title = "Mood & Motivation"
    data4 = Reference(ws, min_col=FIRST_DAY_COL, max_col=LAST_DAY_COL, min_row=MOOD_ROW, max_row=MOTIVATION_ROW)
    cats4 = Reference(ws, min_col=FIRST_DAY_COL, max_col=LAST_DAY_COL, min_row=DAYNUM_ROW)
    chart4.add_data(data4, titles_from_data=False, from_rows=True)
    chart4.set_categories(cats4)
    chart4.series[0].tx = SeriesLabel(v="Mood")
    chart4.series[1].tx = SeriesLabel(v="Motivation")
    chart4.height = 7
    chart4.width = 22
    ws.add_chart(chart4, f"A{LINE_CHART_ROW}")


# ---------------------------------------------------------------------------
# Helpers sheet
# ---------------------------------------------------------------------------
def build_helpers(wb):
    ws = wb.create_sheet("Helpers")
    ws.sheet_state = "hidden"

    ws["A1"] = "Per-habit streak running counts (column = day of month, value = consecutive checks ending that day)"
    ws["A1"].font = MUTED_FONT

    last_day_letter = col(LAST_DAY_COL)

    # Streak helper rows: column A = 0 (start), columns B..AF = day1..day31
    for i in range(N):
        r = HELPER_STREAK_FIRST_ROW + i
        tracker_row = HABIT_FIRST_ROW + i
        ws.cell(row=r, column=1, value=0)
        for d in range(1, DAY_COLS + 1):
            tracker_col = col(FIRST_DAY_COL + d - 1)
            prev_helper_col = col(d)
            cell = ws.cell(row=r, column=d + 1)
            cell.value = f'=IF(Tracker!{tracker_col}{tracker_row}="{CHECK}",{prev_helper_col}{r}+1,0)'

    # Weekly score: Week 1-5 labels + values (referencing Tracker habit rows)
    ws.cell(row=WEEK_LABEL_ROW - 1, column=1, value="Weekly score (Week 1-5)")
    ws.cell(row=WEEK_LABEL_ROW - 1, column=1).font = MUTED_FONT
    for w, (start, end) in enumerate(WEEK_GROUPS, start=1):
        label_cell = ws.cell(row=WEEK_LABEL_ROW, column=w, value=f"Week {w}")
        label_cell.font = MUTED_FONT
        habit_range = f"Tracker!{col(start)}{HABIT_FIRST_ROW}:{col(end)}{HABIT_LAST_ROW}"
        value_cell = ws.cell(row=WEEK_VALUE_ROW, column=w)
        if w <= 4:
            value_cell.value = f'=SUMPRODUCT(--({habit_range}="{CHECK}"))/({N}*7)'
        else:
            value_cell.value = (
                f'=IF(Tracker!$B$6<=28,0,'
                f'SUMPRODUCT(--({habit_range}="{CHECK}"))/({N}*(Tracker!$B$6-28)))'
            )
        value_cell.number_format = "0%"

    # Donut data: Completed / Remaining
    ws.cell(row=DONUT_FIRST_ROW - 1, column=1, value="Overall status data")
    ws.cell(row=DONUT_FIRST_ROW - 1, column=1).font = MUTED_FONT
    check_range = f"Tracker!{col(FIRST_DAY_COL)}{HABIT_FIRST_ROW}:{last_day_letter}{HABIT_LAST_ROW}"
    ws.cell(row=DONUT_FIRST_ROW, column=1, value="Completed")
    ws.cell(row=DONUT_FIRST_ROW, column=2, value=f'=SUMPRODUCT(--({check_range}="{CHECK}"))')
    ws.cell(row=DONUT_FIRST_ROW + 1, column=1, value="Remaining")
    ws.cell(row=DONUT_FIRST_ROW + 1, column=2, value=f'={N}*Tracker!$B$6-B{DONUT_FIRST_ROW}')

    # Ranking tie-break values
    ws.cell(row=TIEBREAK_FIRST_ROW - 1, column=1, value="Ranking tie-break values")
    ws.cell(row=TIEBREAK_FIRST_ROW - 1, column=1).font = MUTED_FONT
    progress_col = col(ANALYSIS_FIRST + 3)
    for i in range(N):
        r = TIEBREAK_FIRST_ROW + i
        tracker_row = HABIT_FIRST_ROW + i
        ws.cell(row=r, column=1, value=f"=Tracker!A{tracker_row}")
        ws.cell(row=r, column=2, value=f"=Tracker!{progress_col}{tracker_row}-ROW()*0.0000001")

    return ws


def main():
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    wb.properties.creator = "Nikolas Moore"
    wb.properties.lastModifiedBy = "Nikolas Moore"

    build_helpers(wb)
    build_tracker(wb)

    wb.active = wb.sheetnames.index("Tracker")

    out = "Goal Tracker Template.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
